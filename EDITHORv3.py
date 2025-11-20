import pdfplumber
import re
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook
import os
import json
import sys

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- FENÊTRE DE L'ÉDITEUR D'EAN (NOUVEAU) ---
class EanEditorWindow(tk.Toplevel):
    def __init__(self, master, corrections_file):
        super().__init__(master)
        self.transient(master)
        self.title("Éditeur de Corrections EAN")
        self.geometry("500x550")
        self.configure(bg="#f2f2f7")
        self.resizable(False, False)
       
        self.corrections_file = corrections_file
        self.corrections = self.load_corrections()

        self.setup_widgets()
        self.populate_listbox()
       
        self.protocol("WM_DELETE_WINDOW", self.save_and_close)
        self.grab_set() # Met la fenêtre au premier plan

    def setup_widgets(self):
        main_frame = tk.Frame(self, bg="#f2f2f7")
        main_frame.pack(padx=15, pady=15, fill="both", expand=True)

        # --- Liste des corrections ---
        tk.Label(main_frame, text="Liste des corrections (Ancien EAN -> Nouveau EAN) :", bg="#f2f2f7", font=("Helvetica Neue", 11, "bold")).pack(anchor="w")
        list_frame = tk.Frame(main_frame)
        list_frame.pack(pady=5, fill="x")
        self.listbox = tk.Listbox(list_frame, height=10, font=("Helvetica Neue", 10))
        self.listbox.pack(side="left", fill="x", expand=True)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.listbox.config(yscrollcommand=scrollbar.set)
        self.listbox.bind("<<ListboxSelect>>", self.on_listbox_select)

        # --- Champs de saisie ---
        tk.Label(main_frame, text="Ancien EAN (lu sur le PDF) :", bg="#f2f2f7", font=("Helvetica Neue", 10)).pack(anchor="w", pady=(10, 0))
        self.old_ean_entry = tk.Entry(main_frame, font=("Helvetica Neue", 10))
        self.old_ean_entry.pack(fill="x", pady=5)

        tk.Label(main_frame, text="Nouveau EAN (le code correct) :", bg="#f2f2f7", font=("Helvetica Neue", 10)).pack(anchor="w", pady=(10, 0))
        self.new_ean_entry = tk.Entry(main_frame, font=("Helvetica Neue", 10))
        self.new_ean_entry.pack(fill="x", pady=5)

        # --- Boutons d'action ---
        button_frame = tk.Frame(main_frame, bg="#f2f2f7")
        button_frame.pack(pady=15, fill="x")
        tk.Button(button_frame, text="Ajouter", command=self.add_correction).pack(side="left", padx=5, expand=True)
        tk.Button(button_frame, text="Modifier", command=self.update_correction).pack(side="left", padx=5, expand=True)
        tk.Button(button_frame, text="Supprimer", command=self.delete_correction).pack(side="left", padx=5, expand=True)

        # --- Bouton de fermeture ---
        tk.Button(main_frame, text="Fermer et Sauvegarder", command=self.save_and_close, bg="#007aff", fg="white", font=("Helvetica Neue", 11, "bold")).pack(side="bottom", fill="x", pady=10)

    def load_corrections(self):
        try:
            with open(self.corrections_file, 'r') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return {}

    def populate_listbox(self):
        self.listbox.delete(0, tk.END)
        for old_ean, new_ean in sorted(self.corrections.items()):
            self.listbox.insert(tk.END, f"{old_ean}  ->  {new_ean}")

    def on_listbox_select(self, event=None):
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            return
       
        selected_text = self.listbox.get(selection_indices[0])
        old_ean, new_ean = [x.strip() for x in selected_text.split("->")]
       
        self.old_ean_entry.delete(0, tk.END)
        self.old_ean_entry.insert(0, old_ean)
        self.new_ean_entry.delete(0, tk.END)
        self.new_ean_entry.insert(0, new_ean)

    def add_correction(self):
        old_ean = self.old_ean_entry.get().strip()
        new_ean = self.new_ean_entry.get().strip()
        if not old_ean or not new_ean:
            messagebox.showwarning("Erreur", "Les deux champs EAN doivent être remplis.", parent=self)
            return
        if not old_ean.isdigit() or not new_ean.isdigit():
            messagebox.showwarning("Erreur", "Les EAN ne doivent contenir que des chiffres.", parent=self)
            return

        self.corrections[old_ean] = new_ean
        self.populate_listbox()
        self.old_ean_entry.delete(0, tk.END)
        self.new_ean_entry.delete(0, tk.END)
        messagebox.showinfo("Succès", "Correction ajoutée.", parent=self)

    def update_correction(self):
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une correction à modifier.", parent=self)
            return
       
        selected_text = self.listbox.get(selection_indices[0])
        original_old_ean = selected_text.split("->")[0].strip()
       
        new_old_ean = self.old_ean_entry.get().strip()
        new_new_ean = self.new_ean_entry.get().strip()
       
        if not new_old_ean or not new_new_ean:
            messagebox.showwarning("Erreur", "Les deux champs EAN ne doivent être vides.", parent=self)
            return

        # Supprimer l'ancienne clé et ajouter la nouvelle
        if original_old_ean in self.corrections:
            del self.corrections[original_old_ean]
        self.corrections[new_old_ean] = new_new_ean
        self.populate_listbox()
        messagebox.showinfo("Succès", "Correction modifiée.", parent=self)
       
    def delete_correction(self):
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("Erreur", "Veuillez sélectionner une correction à supprimer.", parent=self)
            return
       
        selected_text = self.listbox.get(selection_indices[0])
        old_ean_to_delete = selected_text.split("->")[0].strip()
       
        if old_ean_to_delete in self.corrections:
            del self.corrections[old_ean_to_delete]
            self.populate_listbox()
            self.old_ean_entry.delete(0, tk.END)
            self.new_ean_entry.delete(0, tk.END)
            messagebox.showinfo("Succès", "Correction supprimée.", parent=self)

    def save_and_close(self):
        try:
            with open(self.corrections_file, 'w') as f:
                json.dump(self.corrections, f, indent=4)
            self.destroy()
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d’enregistrer les modifications :\n{e}", parent=self)


# --- APPLICATION PRINCIPALE EDITHOR ---
class EdithorApp:
    CONFIG_FILE = 'config.json'
    EAN_CORRECTIONS_FILE = 'corrections_ean.json'

    def __init__(self, root):
        self.root = root
        # ... (le reste du __init__ est identique)
        self.root.title("EDITHOR")
        self.root.geometry("600x780")
        self.root.configure(bg="#f2f2f7")
        self.root.resizable(False, False)
        self.default_base_path = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
        self.excel_template_path = os.path.join(self.default_base_path, "EDI.xlsx")
        from pathlib import Path
        documents_path = Path.home() / "Documents"
        self.output_excel_path = os.path.join(documents_path, "Commande EXCEL EDITHOR")
        os.makedirs(self.output_excel_path, exist_ok=True)
        self.pdf_path = None
        self.load_config()
        self.setup_ui()
   
    def setup_ui(self):
        # ... (le début du setup_ui est identique)
        font_main = ("Helvetica Neue", 11)
        button_style = {
            "font": font_main, "width": 42, "height": 2, "bg": "#007aff",
            "fg": "white", "bd": 0, "activebackground": "#005bb5", "cursor": "hand2"
        }
        try:
            logo = tk.PhotoImage(file=resource_path("EDITHOR2.png"))
            logo_label = tk.Label(self.root, image=logo, bg="#f2f2f7")
            logo_label.image = logo
            logo_label.pack(pady=(20, 10))
        except:
            pass

        tk.Label(self.root, text="Bienvenue dans EDITHOR", font=("Helvetica Neue", 18, "bold"), bg="#f2f2f7", fg="#1c1c1e").pack(pady=(5, 15))
       
        tk.Button(self.root, text="📂 Commencer", command=self.start_process, **button_style).pack(pady=4)
        tk.Button(self.root, text="🔄 Changer le fichier Excel", command=self.change_excel_template_path, **button_style).pack(pady=4)
        tk.Button(self.root, text="📁 Changer le dossier de sortie", command=self.change_output_path, **button_style).pack(pady=4)
       
        # MODIFIÉ : Le bouton appelle maintenant la nouvelle fenêtre
        tk.Button(self.root, text="✍ Gérer les Corrections EAN", command=self.open_ean_editor, **button_style).pack(pady=4)
       
        tk.Button(self.root, text="🧹 Vider le dossier de sortie", command=self.clear_output_path, **button_style).pack(pady=4)
        tk.Button(self.root, text="📂 Ouvrir dossier de sortie", command=self.open_output_folder, **button_style).pack(pady=4)
        tk.Button(self.root, text="📂 Ouvrir dossier EDI", command=self.open_edi_folder, **button_style).pack(pady=4)
       
        footer_frame = tk.Frame(self.root, bg="#f2f2f7")
        footer_frame.pack(expand=True, fill="both")
        tk.Label(footer_frame, text="★★★★★", font=("Helvetica Neue", 12), bg="#f2f2f7", fg="#ffaa00").pack(pady=(10, 0))
        tk.Label(footer_frame, text="Powered by IC - 2025", font=("Helvetica Neue", 9, "italic"), bg="#f2f2f7", fg="#8e8e93").pack(pady=(2, 10))

    # NOUVEAU : Ouvre la fenêtre d'édition
    def open_ean_editor(self):
        EanEditorWindow(self.root, self.EAN_CORRECTIONS_FILE)

    def load_ean_corrections(self):
        try:
            with open(self.EAN_CORRECTIONS_FILE, 'r') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            # S'il n'existe pas ou est corrompu, on en crée un vide
            with open(self.EAN_CORRECTIONS_FILE, 'w') as f:
                json.dump({}, f)
            return {}

    # Le reste du code (start_process, analyser_produit, etc.) est identique à la version précédente
    # qui utilisait déjà le dictionnaire de corrections. Il n'y a rien à changer dans ces fonctions.
   
    def select_pdf_file(self):
        self.pdf_path = filedialog.askopenfilename(title="Sélectionnez le fichier PDF", filetypes=[("PDF files", "*.pdf")])
        if not self.pdf_path:
            messagebox.showerror("Erreur", "Aucun fichier PDF sélectionné.")
           
    def change_excel_template_path(self):
        path = filedialog.askopenfilename(title="Sélectionnez le modèle Excel", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_template_path = path
            self.save_config()

    def change_output_path(self):
        path = filedialog.askdirectory(title="Sélectionnez le dossier de sortie")
        if path:
            self.output_excel_path = path
            self.save_config()

    def clear_output_path(self):
        if self.output_excel_path:
            for f in os.listdir(self.output_excel_path):
                file_path = os.path.join(self.output_excel_path, f)
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            messagebox.showinfo("Succès", "Le dossier de sortie a été vidé.")
           
    def open_output_folder(self):
        if os.path.exists(self.output_excel_path):
            os.startfile(self.output_excel_path)

    def open_edi_folder(self):
        edi_path = os.path.dirname(self.excel_template_path)
        if os.path.exists(edi_path):
            os.startfile(edi_path)

    def start_process(self):
        self.select_pdf_file()
        if not self.pdf_path or not self.excel_template_path or not self.output_excel_path:
            messagebox.showerror("Erreur", "Veuillez définir tous les chemins.")
            return

        ean_corrections = self.load_ean_corrections()
        commandes = self.extraire_et_structurer_texte_pdf(self.pdf_path, ean_corrections)
       
        self.creer_excel_a_partir_du_modele(self.excel_template_path, self.output_excel_path, commandes)
        messagebox.showinfo("Succès", f"Les fichiers Excel ont été créés dans :\n{self.output_excel_path}")

    def extraire_et_structurer_texte_pdf(self, pdf_path, ean_corrections):
        commandes, current_commande, produits, inside_commande = [], None, [], False
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    commandes, current_commande, produits, inside_commande = self.parse_and_structure_text(
                        text, commandes, current_commande, produits, inside_commande, ean_corrections
                    )
        if current_commande and produits:
            current_commande['Produits'] = produits
            commandes.append(current_commande)
        return commandes

    def parse_and_structure_text(self, text, commandes, current_commande, produits, inside_commande, ean_corrections):
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith("Commande n°"):
                inside_commande = True
                if current_commande:
                    current_commande['Produits'] = produits
                    commandes.append(current_commande)
                    produits = []
                current_commande = {}

            if inside_commande:
                if line.startswith("Commande n°"):
                    current_commande['Commande'] = line.split("Commande n°")[1].strip()
                elif line.startswith("Fournisseur"):
                    current_commande['Fournisseur'] = line.split(":")[1].strip()
                elif line.startswith("Document"):
                    current_commande['DateCommande'] = line.split(":")[1].strip()
                elif line.startswith("Livraison le"):
                    current_commande['DateLivraison'] = line.split(":")[1].strip()
                elif "BAK FRANCE" in line:
                    current_commande['NomClient'] = line.split("BAK FRANCE")[1].strip()
                elif line.startswith("Lieu dit"):
                    current_commande['Adresse'] = line
                elif line.startswith("Poids total brut produits"):
                    current_commande['PoidsTotal'] = line.split(":")[1].strip()
                elif line.startswith("Montant total ht commande"):
                    current_commande['MontantTotal'] = line.split(":")[1].strip()
                elif re.match(r"^\d+ \d+", line):
                    produit = self.analyser_produit(line, ean_corrections)
                    if produit:
                        produits.append(produit)
                elif line.startswith("Récapitulatif"):
                    inside_commande = False
                    if current_commande:
                        current_commande['Produits'] = produits
                        commandes.append(current_commande)
                        produits = []
                        current_commande = None
        return commandes, current_commande, produits, inside_commande

    def analyser_produit(self, line, ean_corrections):
        parts = re.split(r'\s+', line)
        if len(parts) >= 6:
            ean_brut = parts[2]
            ean_corrige = ean_corrections.get(ean_brut, ean_brut)
           
            return {
                "EAN": ean_corrige,
                "Description": " ".join(parts[3:-3]),
                "QuantiteCommandee": parts[-3],
                "PCB": parts[-2]
            }
        return {}

    def creer_excel_a_partir_du_modele(self, modele_path, output_path, commandes):
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        for commande in commandes:
            if not commande.get('Produits'):
                continue
            wb = load_workbook(modele_path)
            ws = wb.active
            ws['E2'] = commande.get('DateCommande', "")[:10]
            ws['F2'] = commande.get('DateLivraison', "")[:10]
            nom_client = commande.get('NomClient', "").split("BAK")[0].strip()
            ws['I2'] = ws['K2'] = nom_client
            ws['L2'] = ws['M2'] = ws['N2'] = ""
            numero_commande = commande.get('Commande', "")
            nom_fichier = f"{nom_client}_{numero_commande}".replace(" ", "_").strip("_")
            ws['O2'] = nom_fichier
            for i, produit in enumerate(commande['Produits'], start=4):
                ws[f'C{i}'] = produit.get('EAN', "")
                ws[f'D{i}'] = 'PCE'
                ws[f'E{i}'] = produit.get('Description', "")
                ws[f'F{i}'] = produit.get('QuantiteCommandee', "")
                ws[f'G{i}'] = produit.get('PCB', "")
            wb.save(os.path.join(output_path, f"{nom_fichier}.xlsx"))

    def load_config(self):
        if os.path.exists(self.CONFIG_FILE):
            with open(self.CONFIG_FILE, 'r') as f:
                config = json.load(f)
                self.excel_template_path = config.get('excel_template_path', self.excel_template_path)
                self.output_excel_path = config.get('output_excel_path', self.output_excel_path)
        else:
            self.save_config()

    def save_config(self):
        config = {
            'excel_template_path': self.excel_template_path,
            'output_excel_path': self.output_excel_path
        }
        with open(self.CONFIG_FILE, 'w') as f:
            json.dump(config, f)


def main():
    root = tk.Tk()
    app = EdithorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
